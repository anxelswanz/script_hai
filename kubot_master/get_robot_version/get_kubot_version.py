import json
import os
import time
import paramiko
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import socket
from datetime import datetime
import logging
import concurrent.futures
from threading import Lock
import queue

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class DeviceInfoCollector:
    def __init__(self, username='kubot', password='HairouKubot_@2018!', max_workers=20):
        self.username = username
        self.password = password
        self.timeout = 10  # SSH连接超时时间
        self.max_workers = max_workers  # 最大并发线程数
        self.lock = Lock()  # 线程锁，用于保护共享资源

    def connect_device(self, ip):
        """连接设备"""
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(ip, username=self.username, password=self.password,
                        timeout=self.timeout, look_for_keys=False)
            logger.info(f"成功连接到设备 {ip}")
            return ssh, None
        except paramiko.AuthenticationException:
            error_msg = f"认证失败 - {ip}"
            logger.error(error_msg)
            return None, error_msg
        except paramiko.SSHException as e:
            error_msg = f"SSH连接错误 - {ip}: {str(e)}"
            logger.error(error_msg)
            return None, error_msg
        except socket.timeout:
            error_msg = f"连接超时 - {ip}"
            logger.error(error_msg)
            return None, error_msg
        except Exception as e:
            error_msg = f"连接设备 {ip} 时发生错误: {str(e)}"
            logger.error(error_msg)
            return None, error_msg

    def execute_command(self, ssh, command):
        """执行命令并返回结果"""
        try:
            stdin, stdout, stderr = ssh.exec_command(command, timeout=15)  # 命令执行超时
            output = stdout.read().decode('utf-8').strip()
            error = stderr.read().decode('utf-8').strip()

            if error and "No such file or directory" not in error:
                logger.warning(f"命令执行警告 ({command}): {error}")

            return output, None
        except socket.timeout:
            error_msg = f"命令执行超时: {command}"
            logger.error(error_msg)
            return None, error_msg
        except Exception as e:
            error_msg = f"执行命令失败: {str(e)}"
            logger.error(error_msg)
            return None, error_msg

    def read_remote_file(self, ssh, file_path):
        """读取远程文件内容"""
        try:
            sftp = ssh.open_sftp()
            sftp.sock.settimeout(15)  # SFTP操作超时
            with sftp.file(file_path, 'r') as f:
                content = f.read().decode('utf-8')
            return content, None
        except Exception as e:
            error_msg = f"读取文件 {file_path} 失败: {str(e)}"
            logger.error(error_msg)
            return None, error_msg

    def get_device_version(self, ssh):
        """获取设备版本号"""
        command = 'strings /home/kubot/app/kubot_master_node 2>/dev/null | grep HAIPICKG4'
        output, error = self.execute_command(ssh, command)

        if error:
            return None, error

        if output:
            # 提取版本号，假设输出格式为 "HAIPICKG4_X.X.X"
            lines = output.split('\n')
            if lines:
                version_line = lines[0].strip()
                return version_line, None

        return None, "未找到版本信息"

    def get_feature_model(self, ssh):
        """获取设备型号"""
        file_path = '/home/kubot/app/config__hardware-design.json'
        content, error = self.read_remote_file(ssh, file_path)

        if error:
            return None, error

        try:
            config_data = json.loads(content)
            profiles = config_data.get('profiles', [])
            for profile in profiles:
                profile_data = profile.get('profile-data', {})
                if 'feature_model' in profile_data:
                    return profile_data['feature_model'], None
            return None, "未找到feature_model字段"
        except json.JSONDecodeError as e:
            return None, f"JSON解析错误: {str(e)}"
        except Exception as e:
            return None, f"解析配置文件失败: {str(e)}"

    def get_robot_id(self, ssh):
        """获取设备ID"""
        file_path = '/home/kubot/app/config__system-deployment.json'
        content, error = self.read_remote_file(ssh, file_path)

        if error:
            return None, error

        try:
            config_data = json.loads(content)
            profiles = config_data.get('profiles', [])
            for profile in profiles:
                profile_data = profile.get('profile-data', {})
                if 'robot_id' in profile_data:
                    return str(profile_data['robot_id']), None
            return None, "未找到robot_id字段"
        except json.JSONDecodeError as e:
            return None, f"JSON解析错误: {str(e)}"
        except Exception as e:
            return None, f"解析配置文件失败: {str(e)}"

    def get_device_info_single(self, ip):
        """获取单个设备的完整信息（供线程调用）"""
        logger.info(f"开始收集设备 {ip} 的信息...")

        device_info = {
            'ip': ip,
            'robot_id': 'N/A',
            'feature_model': 'N/A',
            'version': 'N/A',
            'status': '成功',
            'error_message': '',
            'collection_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        ssh, error = self.connect_device(ip)
        if error:
            device_info['status'] = '失败'
            device_info['error_message'] = error
            return device_info

        try:
            # 并行执行三个信息获取任务
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                version_future = executor.submit(self.get_device_version, ssh)
                feature_model_future = executor.submit(self.get_feature_model, ssh)
                robot_id_future = executor.submit(self.get_robot_id, ssh)

                # 等待所有任务完成
                version, version_error = version_future.result(timeout=20)
                feature_model, feature_model_error = feature_model_future.result(timeout=20)
                robot_id, robot_id_error = robot_id_future.result(timeout=20)

            # 处理结果
            if version_error:
                device_info['error_message'] += f"版本获取失败: {version_error}; "
            else:
                device_info['version'] = version

            if feature_model_error:
                device_info['error_message'] += f"型号获取失败: {feature_model_error}; "
            else:
                device_info['feature_model'] = feature_model

            if robot_id_error:
                device_info['error_message'] += f"ID获取失败: {robot_id_error}; "
            else:
                device_info['robot_id'] = robot_id

            # 清理错误信息
            if device_info['error_message']:
                device_info['status'] = '部分成功'
                device_info['error_message'] = device_info['error_message'].rstrip('; ')

            logger.info(f"设备 {ip} 信息收集完成 - 状态: {device_info['status']}")

        except concurrent.futures.TimeoutError:
            device_info['status'] = '失败'
            device_info['error_message'] = "信息收集超时"
            logger.error(f"设备 {ip} 信息收集超时")
        except Exception as e:
            device_info['status'] = '失败'
            device_info['error_message'] = f"收集过程中发生异常: {str(e)}"
            logger.error(f"设备 {ip} 信息收集异常: {str(e)}")
        finally:
            ssh.close()

        return device_info

    def get_all_devices_info(self, device_ips):
        """并行获取所有设备信息"""
        devices_info = []

        logger.info(f"开始并行收集 {len(device_ips)} 个设备的信息，并发数: {self.max_workers}")

        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_ip = {executor.submit(self.get_device_info_single, ip): ip for ip in device_ips}

            # 处理完成的任务
            for future in concurrent.futures.as_completed(future_to_ip):
                ip = future_to_ip[future]
                try:
                    device_info = future.result()
                    with self.lock:
                        devices_info.append(device_info)
                except Exception as exc:
                    logger.error(f"设备 {ip} 生成异常: {exc}")
                    with self.lock:
                        devices_info.append({
                            'ip': ip,
                            'robot_id': 'N/A',
                            'feature_model': 'N/A',
                            'version': 'N/A',
                            'status': '失败',
                            'error_message': f'任务执行异常: {exc}',
                            'collection_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        })

        return devices_info


def load_device_ips(filename='devices.json'):
    """加载设备IP列表"""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            device_ips = json.load(f)
        logger.info(f"成功加载 {len(device_ips)} 个设备IP")
        return device_ips
    except FileNotFoundError:
        logger.error(f"设备列表文件 {filename} 未找到")
        return []
    except json.JSONDecodeError as e:
        logger.error(f"设备列表文件格式错误: {str(e)}")
        return []


def create_excel_report(devices_info):
    """创建Excel报告"""
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'设备信息报告_{timestamp}.xlsx'

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "设备信息"

    # 设置表头
    headers = ['IP地址', '设备ID', '设备型号', '设备版本号', '收集状态', '错误信息', '收集时间']
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 添加数据并设置状态颜色
    status_colors = {
        '成功': '00FF00',  # 绿色
        '部分成功': 'FFFF00',  # 黄色
        '失败': 'FF0000'  # 红色
    }

    for device in devices_info:
        row = [
            device['ip'],
            device['robot_id'],
            device['feature_model'],
            device['version'],
            device['status'],
            device['error_message'],
            device['collection_time']
        ]
        ws.append(row)

        # 为状态列设置颜色
        status_cell = ws.cell(row=ws.max_row, column=5)
        if device['status'] in status_colors:
            status_cell.fill = PatternFill(start_color=status_colors[device['status']],
                                           end_color=status_colors[device['status']],
                                           fill_type='solid')

    # 设置列宽
    column_widths = [15, 10, 15, 20, 12, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 保存文件
    wb.save(filename)
    logger.info(f"Excel报告已生成: {filename}")
    return filename


def print_statistics(devices_info):
    """打印统计信息"""
    total = len(devices_info)
    success_count = sum(1 for d in devices_info if d['status'] == '成功')
    partial_success_count = sum(1 for d in devices_info if d['status'] == '部分成功')
    fail_count = sum(1 for d in devices_info if d['status'] == '失败')

    logger.info("=" * 50)
    logger.info("设备信息收集任务完成")
    logger.info(f"总计设备: {total}")
    logger.info(f"成功: {success_count} ({success_count / total * 100:.1f}%)")
    logger.info(f"部分成功: {partial_success_count} ({partial_success_count / total * 100:.1f}%)")
    logger.info(f"失败: {fail_count} ({fail_count / total * 100:.1f}%)")
    logger.info("=" * 50)


def main():
    """主函数"""
    logger.info("开始设备信息收集任务")

    # 检查当前目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    logger.info(f"工作目录: {current_dir}")

    # 加载设备IP列表
    device_ips = load_device_ips()
    if not device_ips:
        logger.error("没有找到可用的设备IP，程序退出")
        return

    # 创建收集器，设置并发数（可根据网络情况调整）
    collector = DeviceInfoCollector(max_workers=20)  # 可同时处理20台设备

    # 记录开始时间
    start_time = time.time()

    # 并行收集所有设备信息
    devices_info = collector.get_all_devices_info(device_ips)

    # 记录结束时间
    end_time = time.time()

    # 生成报告
    report_filename = create_excel_report(devices_info)

    # 输出统计信息
    print_statistics(devices_info)
    logger.info(f"总耗时: {end_time - start_time:.2f} 秒")
    logger.info(f"平均每台设备: {(end_time - start_time) / len(device_ips):.2f} 秒")
    logger.info(f"报告文件: {report_filename}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logger.info("程序被用户中断")
    except Exception as e:
        logger.error(f"程序执行异常: {str(e)}")